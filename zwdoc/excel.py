from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook

class Excel(object):
    """Excel Class

    .. code-block:: Python

        wb = Excel('my.xlsx')
        wb.sheetnames
        ws = wb[0]
        ws = wb['sheetA']

        # add whole row
        * append(['This is A1', 'This is B1', 'This is C1'])
        * **or** append({'A' : 'This is A1', 'C' : 'This is C1'})
        * **or** append({1 : 'This is A1', 3 : 'This is C1'})

        ws['A4']
        cell_range = ws['A1':'C2']
        coll_range = ws['C:D']
        rows_range = ws[5:10]
        tuple(ws.rows), tuple(ws.columns)

    """
    path = property(lambda o: o._path, lambda o, v: setattr(o, '_path', v))
    active_sheet = property(lambda o: o.wb.active)
    sheet_count = property(lambda o: len(o.wb._sheets))
    sheets = property(lambda o: [Sheet(a) for a in o.wb.worksheets])
    readonly = property(lambda o: o._readonly, lambda o, v: setattr(o, '_readonly', v))
    overwrite = property(lambda o: o._overwrite, lambda o, v: setattr(o, '_overwrite', v))

    def __init__(self, pth=None, overwrite=False) -> None:
        self._path = Path(pth) if pth else None
        self._readonly = False if (pth is None or not self._path.exists()) else True
        self._readonly = self._readonly if overwrite is False else False
        self._overwrite = overwrite
        self.ext = Path(pth).suffix if pth else None
        self.wb = None
    
    def __enter__(self) -> Excel:
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.save()
    
    def open(self) -> Excel:
        if self.wb:
            return self
        if self.path and self.path.exists():
            self.wb = load_workbook(str(self.path))
        else:
            self.wb = Workbook()
        return self

    def save(self) -> Excel:
        if self.path is None or self.readonly:
            return self
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.overwrite and self.path.exists():
            self.path.unlink()
        self.wb.save(self.path)
        return self
    
    def create_sheet(self, title = None, index = None) -> None:
        wb = self.wb
        title = title or f'New Sheet {len(wb._sheets)+1}'
        if index is None:
            wb.create_sheet(title)
        else:
            wb.create_sheet(title, index)
    
    def __getitem__(self, key):
        wb = self.wb
        if isinstance(key, int):
            idx = key
            names = wb.sheetnames
            if idx > len(names) - 1:
                raise KeyError('Worksheet {0} does not exist.'.format(key))
            return Sheet(wb[names[idx]])
        return Sheet(wb[key])
    
    def __getattr__(self, name):
        return getattr(self.wb, name)

class Sheet(object):
    def __init__(self, o) -> None:
        self._o = o

    def __getattr__(self, name):
        return getattr(self._o, name)